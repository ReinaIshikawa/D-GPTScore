import pandas as pd
import random
import itertools
import os 
import numpy as np
from openpyxl import load_workbook

import torch
import sys
sys.path.append(os.path.abspath(os.getcwd()))
from evaluation.utils.eval_utils import get_csv_path

def seed_everything(seed: int):
    random.seed(seed)
    os.environ['PYTHONHASHSEED'] = str(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = True


def user_study_reader(users, gen_method_list, data_dir) -> pd.DataFrame:
    df: pd.DataFrame | None = None
    for user in users:
        output_path = os.path.join(data_dir, f"{user}/userStudy_{user}.xlsx")
        # output_path = os.path.join(data_dir, f"{user}/userStudy.xlsx")
        wb = load_workbook(output_path)
        ws1 = wb['Sheet1']
        ws2 = wb['Sheet2']
        row_counter = 4

        i = 0
        to_save_data = []
        while(True):
            i+=1
            info = ws2.cell(row=row_counter, column=1).value
            mode, prompt_type, id_ =info.split("_") # type: ignore
            user_study_data = {}
            for sid in range(6):
                score = ws1.cell(row=row_counter+1, column=2+sid).value
                method_name = ws2.cell(row=row_counter, column=2+sid).value
                if score not in [1,2,3,4,5,6,7,8,9,10]:
                    print(f"user{user}: {mode}_{prompt_type}_{id_} cell is Nan")
                    score = 5
                user_study_data[method_name]=score
                

            for sd in gen_method_list:
                to_save_data.append({
                    "method":sd,
                    "mode":mode,
                    "prompt_type":prompt_type,
                    "id_":id_,
                    user:user_study_data[sd],
            })
            row_counter += 4
            if i >= 120:
                break
        usr_df = pd.DataFrame(to_save_data)
        if df is None:
            df = usr_df
        else:
            df = pd.merge(
                df, 
                usr_df,
                on=['method', 'mode', 'prompt_type', 'id_'], 
                how='outer')
    if df is None:
        raise ValueError("No data was processed")
    return df


def get_user_study_target() -> list[tuple[str, str, int]]:
    #return list of tuple (mode, prompt_type, id_)
    user_study_target = []
    m_list = ["easy", "medium","hard"]
    p_list = ["simple", "action+layout", "action+expression", "action+background", "all"]
    for mode in m_list:
        if mode == "easy":
            seed_ = 111
            index_list = list(range(52))
        if mode == "medium":
            seed_ = 123
            index_list = list(range(52))
        if mode == "hard":
            seed_ = 234
            index_list = list(range(92))
        random.seed(seed_)
        random.shuffle(index_list)
        index_list = index_list[:40]
        user_study_target += list(zip(itertools.cycle([mode]), itertools.cycle(p_list), index_list))
    return user_study_target


def user_df_splitter(
    method_list: list[str], 
    user_df: pd.DataFrame, 
    mode_list: list[str], 
    prompt_type_list: list[str],
    users: list[str], 
) -> pd.DataFrame:

    user_study_target = get_user_study_target()

    combined=pd.DataFrame()
    for method in method_list:
        for mode, prompt_type, id_ in user_study_target:
            if mode in mode_list and prompt_type in prompt_type_list:
                target_df = user_df[
                    (user_df['method'] == method) & (user_df['mode'] == mode) & (user_df['prompt_type'] == prompt_type) & (user_df['id_'] == str(id_))
                ]
                target_df = target_df[users]
                if isinstance(target_df, pd.Series):
                    target_df = target_df.to_frame().T
                elif isinstance(target_df, np.ndarray):
                    target_df = pd.DataFrame(target_df, columns=pd.Index(users))
                if combined.empty:
                    combined = target_df
                else:
                    combined = pd.concat([combined, target_df], axis=0, ignore_index=True) 

    return combined

def get_arcface_single_score_df(metric_df: pd.DataFrame) -> pd.DataFrame:
    df_single = metric_df[(metric_df["mode"] == "easy")].copy()
    df_single['score'] = df_single['p1_sim_score']
    df_multi = metric_df[(metric_df["mode"] == "medium") | (metric_df["mode"] == "hard")].copy()
    df_multi['score'] = df_multi[['p1_sim_score', 'p2_sim_score']].mean(1)
    return pd.concat([df_single,df_multi], axis=0) # type: ignore

def metric_df_of_user_study(
    gen_method_list: list[str], 
    mode_list: list[str], 
    prompt_type_list: list[str], 
    results_dir: str, 
    metric_name: str, 
    header: str="") -> pd.DataFrame:
    """
    get gpt dataframe of which is used for user study
    """

    user_study_target = get_user_study_target()

    combined=pd.DataFrame()
    if metric_name == "GPT_ours":
        scores = [f"score{i}" for i in range(1, 19)]
    else:
        scores = ["score"]

    for gen_method in gen_method_list:
        output_csv_path = get_csv_path(
            results_dir=results_dir, 
            metric_name=metric_name, # gpt_ours/ArcFace/ ...
            gen_method=gen_method, # 01_CustomDiffusion/02_OMG_lora/ ...
            header=header
        )
        metric_df = pd.read_csv(output_csv_path, on_bad_lines='skip')
        metric_df.fillna(0, inplace=True)

        if metric_name == "ArcFace":
            metric_df = get_arcface_single_score_df(metric_df)

        for mode, prompt_type, id_ in user_study_target:
            if mode in mode_list and prompt_type in prompt_type_list:
                target_df = metric_df[
                    (metric_df['mode'] == mode) & (metric_df['prompt_type'] == prompt_type) & (metric_df['id_'] == id_)
                ]
                target_df = target_df[scores]
                if combined.empty:#type: ignore
                    combined = target_df
                else:
                    combined = pd.concat([combined, target_df], axis=0, ignore_index=True) # type: ignore
    return combined # type: ignore

def metric_df_all(
    gen_method_list: list[str], 
    mode_list: list[str], 
    prompt_type_list: list[str], 
    results_dir: str, 
    metric_name: str, 
    header: str="") -> pd.DataFrame:

    combined=pd.DataFrame()
    if metric_name == "GPT_ours":
        scores = [f"score{i}" for i in range(1, 19)]
    else:
        scores = ["score"]

    for gen_method in gen_method_list:
        output_csv_path = get_csv_path(
            results_dir=results_dir, 
            metric_name=metric_name, # gpt_ours/ArcFace/ ...
            gen_method=gen_method, # 01_CustomDiffusion/02_OMG_lora/ ...
            header=header
        )
        metric_df = pd.read_csv(output_csv_path, on_bad_lines='skip')
        metric_df.fillna(0, inplace=True)

        if metric_name == "ArcFace":
            metric_df = get_arcface_single_score_df(metric_df)

        for mode, prompt_type in itertools.product(mode_list, prompt_type_list):
            target_df = metric_df[
                (metric_df['mode'] == mode) & (metric_df['prompt_type'] == prompt_type) 
            ]
            sorted_df = target_df.sort_values(by='id_') # type: ignore
            sorted_df = sorted_df[scores]
            if combined.empty:
                combined = sorted_df
            else:
                combined = pd.concat([combined, sorted_df], axis=0, ignore_index=True)
    return combined # type: ignore

def metric_target_reader(
    gen_method: str, 
    metric_name: str, 
    results_dir: str, 
    mode: str, 
    prompt_type: str, 
    idx: int, 
    header: str=""
) -> pd.DataFrame:

    if metric_name == "GPT_ours":
        scores = [f"score{i}" for i in range(1, 19)]
    else:
        scores = ["score"]

    gpt_csv_path = get_csv_path(
        results_dir=results_dir, 
        metric_name=metric_name, # gpt_ours/ArcFace/ ...
        gen_method=gen_method, # 01_CustomDiffusion/02_OMG_lora/ ...
        header=header
    )
    gen_df = pd.read_csv(gpt_csv_path, on_bad_lines='skip')
    gen_df.fillna(0, inplace=True)

    if metric_name == "ArcFace":
        gen_df = get_arcface_single_score_df(gen_df)
    target_df = gen_df[
        (gen_df['mode'] == mode) & (gen_df['prompt_type'] == prompt_type) & (gen_df['id_'] == idx)
    ]
    target_df = target_df[scores]
    return target_df # type: ignore


def user_score_target_reader(
    method: str, 
    user_df: pd.DataFrame, 
    users: list[str], 
    mode: str, 
    prompt_type: str, 
    idx: int
) -> pd.DataFrame:
    """
    Get user scores from user_df
    """

    target_df = user_df[
        (user_df['method'] == method) & (user_df['mode'] == mode) & (user_df['prompt_type'] == prompt_type) & (user_df['id_'] == str(idx))
    ]
    target_df = target_df[users]
    return target_df # type: ignore


def gpt_df_tokennum_reader(
    gen_method_list: list[str], 
    gpt_method: str, 
    results_dir: str, 
    header: str=""
) -> pd.DataFrame:

    combined=pd.DataFrame()
    for gen_method in gen_method_list:
        csv_path = get_csv_path(
            results_dir=results_dir, 
            metric_name=gpt_method, # gpt_ours/ArcFace/ ...
            gen_method=gen_method, # 01_CustomDiffusion/02_OMG_lora/ ...
            header=header
        )
        gpt_df = pd.read_csv(csv_path, on_bad_lines='skip')

        gpt_df = gpt_df[["comp_tok_sum", "prompt_tok_sum"]]
        if combined.empty:
            combined = gpt_df
        else:
            combined = pd.concat([combined, gpt_df], axis=0, ignore_index=True)
    return combined.mean(axis=0) # type: ignore




